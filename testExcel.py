import time
import openpyxl

workbook = openpyxl.load_workbook(r"C:\Users\sriye\OneDrive\Documents\Testing-Selenium\test_files\SeleniumPython.xlsx")
sheet = workbook.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Login"
#workbook.save(r"C:\Users\sriye\OneDrive\Documents\Testing-Selenium\test_files\SeleniumPython.xlsx")
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column=2).value == "Login":
         print("Login testcase found at row", i)
    for j in range(1, sheet.max_column + 1):
        print(sheet.cell(row=i, column=j).value, end=" ")
    print()
workbook.save(r"C:\Users\sriye\OneDrive\Documents\Testing-Selenium\test_files\SeleniumPython.xlsx")
