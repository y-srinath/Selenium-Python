import time
from pathlib import Path
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

download_dir = Path(__file__).resolve().parent / "test_files"
download_dir.mkdir(exist_ok=True)
filePath = download_dir / "download.xlsx"

if filePath.exists():
    filePath.unlink()

chrome_options = webdriver.ChromeOptions()
chrome_options.add_experimental_option("prefs", {
    "download.default_directory": str(download_dir),
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True,
})

driver = webdriver.Chrome(options=chrome_options)
driver.implicitly_wait(5) # Set an implicit wait to handle dynamic content loading
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()

WebDriverWait(driver, 10).until(
    lambda _: filePath.exists() and not any(download_dir.glob("*.crdownload"))
)

fruit_name = "Banana"
newValue = "49"

def update_excel_data(filePath, fruit_name, colName, new_value):
    # Edit the Excel file to update the value based on searchTerm and colName
    workbook = openpyxl.load_workbook(filePath)
    worksheet = workbook.active
    col = None
    row = None

    for i in range(1, worksheet.max_column + 1):
        if str(worksheet.cell(row=1, column=i).value).strip().lower() == colName.lower():
            col = i

    for i in range(1, worksheet.max_row + 1):
        if str(worksheet.cell(row=i, column=2).value).strip() == fruit_name:
            row = i

    if row is None:
        raise ValueError(f"Could not find fruit '{fruit_name}' in the Excel file")
    if col is None:
        raise ValueError(f"Could not find column '{colName}' in the Excel file")

    original_value = worksheet.cell(row=row, column=col).value
    worksheet.cell(row=row, column=col).value = new_value
    workbook.save(filePath)
    return original_value

originalValue = update_excel_data(filePath, fruit_name, "price", newValue)
print("Original price of", fruit_name, "was:", originalValue)

def get_column_id(column_name):
    column = driver.find_element(
        By.XPATH,
        f"//*[normalize-space()='{column_name}']/ancestor-or-self::*[@data-column-id][1]"
    )
    return column.get_attribute("data-column-id")

#Upload edited Excel file
file_input = driver.find_element(By.ID, "fileinput")
file_input.send_keys(str(filePath))
toast_message = WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".Toastify__toast-body")))
assert "Updated Excel Data Successfully." in toast_message.text

priceCol = get_column_id("Price")
fruitCol = get_column_id("Fruit Name")
price_locator = (
    By.XPATH,
    f"//div[@role='row'][.//div[@data-column-id='{fruitCol}' and normalize-space()='{fruit_name}']]"
    f"//div[@data-column-id='{priceCol}']"
)
actual_Price = WebDriverWait(driver, 5).until(
    lambda _: driver.find_element(*price_locator).text == newValue
    and driver.find_element(*price_locator).text
)
print("Updated price of", fruit_name, "is:", actual_Price)
assert actual_Price == newValue, "Price not updated correctly in the web application"

time.sleep(3)
driver.quit()
